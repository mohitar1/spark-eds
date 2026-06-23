#!/usr/bin/env python3
"""Generate DA-ready .xlsx config sheets for the spark-eds asset portal.

Two sheets are produced:

  config-access-application.xlsx  ->  /config/access/application
  config-access-users.xlsx        ->  /config/access/users

Upload these DIRECTLY into DA at:
  https://da.live/edit#/mohitar1/spark-eds
under the path:  config/access/

The Worker fetches them as JSON via fetchHelixSheet().
Column names must match exactly what the Worker code reads.

Sheet designs:

/config/access/application
  Keyed by `email` (also accepts domain and `*` wildcard).
  Controls who can log in, who has admin / sudo / preview access.
  Columns: email | permissions

  permissions is a comma-separated list of tokens:
    preview   — can log into preview Worker hostnames
    sudo      — can impersonate other users via SUDO_* cookies
    admin     — full admin: bypasses all asset filters, can see all reports

/config/access/users
  Keyed by `email` (also accepts domain and `*` wildcard).
  Controls per-user / per-domain asset visibility.
  Columns: email | roles | userType | countries

  roles        — comma-separated: admin
  userType     — 'internal' or 'external'
                 Overrides the automatic domain-based classification.
                 Leave blank to use the default (adobe.com → internal, else external).
  countries    — comma-separated ISO 3166-1 alpha-2 codes the user can see assets for.
                 Special value 'global' is always added automatically by the Worker.
                 Leave blank to restrict the user to their own country only (from JWT).

UPLOAD INSTRUCTIONS
  1. Open https://da.live/edit#/mohitar1/spark-eds
  2. Create folder:  config/access/  (if it doesn't exist)
  3. Drag-and-drop both .xlsx files into that folder.
  4. DA publishes them at /config/access/application.json and /config/access/users.json
  5. Cloudflare Workers picks them up on the next request (cached by EDS CDN,
     invalidated on publish).
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).parent

# ── Palette ──────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill('solid', fgColor='1F3864')   # dark navy
SECTION_FILL  = PatternFill('solid', fgColor='D9E1F2')   # light blue-grey
COMMENT_FILL  = PatternFill('solid', fgColor='F2F2F2')   # light grey
HEADER_FONT   = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
LABEL_FONT    = Font(bold=True, name='Calibri', size=10)
NORMAL_FONT   = Font(name='Calibri', size=10)
COMMENT_FONT  = Font(italic=True, color='595959', name='Calibri', size=9)


def style_header(cell, text):
    cell.value = text
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def style_section(cell, text):
    cell.value = text
    cell.font = LABEL_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(horizontal='left', vertical='center')


def style_comment(ws, row, ncols, text):
    cell = ws.cell(row=row, column=1)
    cell.value = f'▸  {text}'
    cell.font = COMMENT_FONT
    cell.fill = COMMENT_FILL
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row, end_column=ncols,
    )
    ws.row_dimensions[row].height = 24


def style_data(cell, value, bold=False):
    cell.value = value
    cell.font = Font(bold=bold, name='Calibri', size=10)
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def freeze_header(ws):
    ws.freeze_panes = 'A2'


# ─────────────────────────────────────────────────────────────────────────────
# /config/access/application
# ─────────────────────────────────────────────────────────────────────────────
def build_application_sheet():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'data'

    COLS = ['email', 'permissions']
    NCOLS = len(COLS)

    # ── Header row ────────────────────────────────────────────────────────────
    for c, col in enumerate(COLS, 1):
        style_header(ws.cell(row=1, column=c), col)
    ws.row_dimensions[1].height = 28
    freeze_header(ws)

    row = 2

    # ── Section: Wildcard (all authenticated users) ───────────────────────────
    style_comment(ws, row, NCOLS,
        'Wildcard (*): permissions granted to every authenticated user. '
        '"preview" lets users log into non-production Worker hostnames.')
    row += 1

    data = [
        ('*',             'preview'),
    ]
    for email, perms in data:
        style_data(ws.cell(row=row, column=1), email)
        style_data(ws.cell(row=row, column=2), perms)
        row += 1

    # ── Section: Internal domain (adobe.com employees) ────────────────────────
    style_comment(ws, row, NCOLS,
        'Domain-level: all @adobe.com addresses inherit these permissions automatically.')
    row += 1

    data = [
        ('adobe.com',     'preview'),
    ]
    for email, perms in data:
        style_data(ws.cell(row=row, column=1), email)
        style_data(ws.cell(row=row, column=2), perms)
        row += 1

    # ── Section: Named admins ─────────────────────────────────────────────────
    style_comment(ws, row, NCOLS,
        'Named admin accounts: preview + sudo (impersonation) + admin (full bypass). '
        'Add any email that should have full access to all assets and all reports.')
    row += 1

    data = [
        ('mohitar@adobe.com',   'preview,sudo,admin'),
    ]
    for email, perms in data:
        style_data(ws.cell(row=row, column=1), email, bold=True)
        style_data(ws.cell(row=row, column=2), perms)
        row += 1

    # ── Section: External demo users ─────────────────────────────────────────
    style_comment(ws, row, NCOLS,
        'External demo accounts: only "preview" so they can log in. '
        'Asset visibility is controlled by /config/access/users, not here.')
    row += 1

    data = [
        ('demo.external.us@frescopa.coffee',   'preview'),
        ('demo.external.es@frescopa.coffee',   'preview'),
        ('demo.external.gb@frescopa.coffee',   'preview'),
        ('demo.agency@frescopa.coffee',        'preview'),
    ]
    for email, perms in data:
        style_data(ws.cell(row=row, column=1), email)
        style_data(ws.cell(row=row, column=2), perms)
        row += 1

    set_col_widths(ws, [38, 30])

    path = OUT / 'application.xlsx'
    wb.save(path)
    print(f'Wrote {path}')
    print('  → upload into DA: mohitar1/spark-eds  at path  config/access/application')


# ─────────────────────────────────────────────────────────────────────────────
# /config/access/users
# ─────────────────────────────────────────────────────────────────────────────
def build_users_sheet():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'data'

    COLS = ['email', 'roles', 'userType', 'countries']
    NCOLS = len(COLS)

    # ── Header row ────────────────────────────────────────────────────────────
    for c, col in enumerate(COLS, 1):
        style_header(ws.cell(row=1, column=c), col)
    ws.row_dimensions[1].height = 28
    freeze_header(ws)

    row = 2

    # ── Section: Wildcard defaults ────────────────────────────────────────────
    style_comment(ws, row, NCOLS,
        'Wildcard (*): baseline for every authenticated user. '
        'userType and countries left blank — computed from JWT claims at runtime. '
        'No roles granted by default.')
    row += 1

    data = [
        #  email   roles  userType  countries
        ('*',      '',    '',       ''),
    ]
    for email, roles, utype, countries in data:
        style_data(ws.cell(row=row, column=1), email)
        style_data(ws.cell(row=row, column=2), roles)
        style_data(ws.cell(row=row, column=3), utype)
        style_data(ws.cell(row=row, column=4), countries)
        row += 1

    # ── Section: Internal domain (adobe.com) ──────────────────────────────────
    style_comment(ws, row, NCOLS,
        'Domain entry: every @adobe.com user is classified as "internal". '
        'countries left blank — each user sees assets for their own country (from JWT ctry claim) '
        'plus global assets automatically.')
    row += 1

    data = [
        ('adobe.com',  '',      'internal',  ''),
    ]
    for email, roles, utype, countries in data:
        style_data(ws.cell(row=row, column=1), email)
        style_data(ws.cell(row=row, column=2), roles)
        style_data(ws.cell(row=row, column=3), utype)
        style_data(ws.cell(row=row, column=4), countries)
        row += 1

    # ── Section: Named admin ──────────────────────────────────────────────────
    style_comment(ws, row, NCOLS,
        'Named admin: roles=admin bypasses all asset filters (buildAssetAuthClauses returns []). '
        'userType and countries not needed — admin sees everything.')
    row += 1

    data = [
        ('mohitar@adobe.com',  'admin',  'internal',  'IN'),
    ]
    for email, roles, utype, countries in data:
        style_data(ws.cell(row=row, column=1), email, bold=True)
        style_data(ws.cell(row=row, column=2), roles, bold=True)
        style_data(ws.cell(row=row, column=3), utype)
        style_data(ws.cell(row=row, column=4), countries)
        row += 1

    # ── Section: Demo external users ──────────────────────────────────────────
    style_comment(ws, row, NCOLS,
        'Demo external users: simulate agency partners / distributors from different countries. '
        'userType=external means they only see assets tagged external or all. '
        'countries restricts them to assets tagged with their market + global.')
    row += 1

    data = [
        # US-based external (e.g. US distributor)
        ('demo.external.us@frescopa.coffee',  '',  'external',  'US'),
        # Spain-based external (e.g. ES agency partner)
        ('demo.external.es@frescopa.coffee',  '',  'external',  'ES'),
        # UK-based external
        ('demo.external.gb@frescopa.coffee',  '',  'external',  'GB'),
        # Agency with multi-market access (EMEA agency seeing ES + GB + global)
        ('demo.agency@frescopa.coffee',       '',  'external',  'ES,GB'),
    ]
    for email, roles, utype, countries in data:
        style_data(ws.cell(row=row, column=1), email)
        style_data(ws.cell(row=row, column=2), roles)
        style_data(ws.cell(row=row, column=3), utype)
        style_data(ws.cell(row=row, column=4), countries)
        row += 1

    # ── Section: Demo internal users (multi-country) ──────────────────────────
    style_comment(ws, row, NCOLS,
        'Demo internal users: simulate Adobe employees with broader market access. '
        'A global brand manager sees all markets; a regional lead sees their region only. '
        'userType left blank because adobe.com domain already resolves to internal.')
    row += 1

    data = [
        # Global brand manager — sees all markets (US + ES + GB + global)
        ('global.manager@adobe.com',   '',  '',  'US,ES,GB'),
        # EMEA regional lead — sees ES + GB + global
        ('emea.lead@adobe.com',        '',  '',  'ES,GB'),
        # US market manager — sees US + global only
        ('us.manager@adobe.com',       '',  '',  'US'),
    ]
    for email, roles, utype, countries in data:
        style_data(ws.cell(row=row, column=1), email)
        style_data(ws.cell(row=row, column=2), roles)
        style_data(ws.cell(row=row, column=3), utype)
        style_data(ws.cell(row=row, column=4), countries)
        row += 1

    # ── Section: Edge-case overrides ──────────────────────────────────────────
    style_comment(ws, row, NCOLS,
        'Override examples: promote a specific non-Adobe email to internal (e.g. embedded '
        'contractor who needs access to internal assets), or demote to external for demo purposes.')
    row += 1

    data = [
        # Contractor on a non-Adobe domain promoted to internal
        ('contractor@partner-agency.com',  '',  'internal',  'US'),
    ]
    for email, roles, utype, countries in data:
        style_data(ws.cell(row=row, column=1), email)
        style_data(ws.cell(row=row, column=2), roles)
        style_data(ws.cell(row=row, column=3), utype)
        style_data(ws.cell(row=row, column=4), countries)
        row += 1

    set_col_widths(ws, [36, 12, 12, 22])

    path = OUT / 'users.xlsx'
    wb.save(path)
    print(f'Wrote {path}')
    print('  → upload into DA: mohitar1/spark-eds  at path  config/access/users')


# ─────────────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    build_application_sheet()
    build_users_sheet()
    print(
        '\nDone.'
        '\n'
        '\nUpload both .xlsx files into DA at:'
        '\n  https://da.live/edit#/mohitar1/spark-eds'
        '\n  folder:  config/access/'
        '\n'
        '\nThe Worker reads them as:'
        '\n  /config/access/application.json  (keyed by email/domain/*)'
        '\n  /config/access/users.json        (keyed by email/domain/*)'
        '\n'
        '\nAfter uploading, click "Publish" on each file in DA.'
        '\nThe Cloudflare Worker picks up the new values on the next request.'
        '\n'
        '\nContent Hub asset tagging required for filters to take effect:'
        '\n  custom:userType  =  internal | external | all'
        '\n  custom:country   =  US | ES | GB | global  (ISO 3166-1 alpha-2 or "global")'
    )
